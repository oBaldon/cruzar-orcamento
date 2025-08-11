# src/cruzar_orcamento/adapters/sinapi.py
from __future__ import annotations

import logging
import re
from typing import Optional, Dict
import unicodedata

import pandas as pd
from openpyxl import load_workbook

from ..models import Item, CanonDict
from ..utils.utils_text import norm_code

logger = logging.getLogger(__name__)

# ----------------- helpers -----------------

def _strip_accents(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()

def _norm(s: str) -> str:
    if not isinstance(s, str):
        s = "" if pd.isna(s) else str(s)
    return _strip_accents(s).lower().strip()

_HYPERLINK_CODE_RE = re.compile(r"(\d+)\)?\s*$")
_DIGIT_CODE_RE = re.compile(r"^\d{3,}$")

def _extract_code_from_formula(formula: str) -> Optional[str]:
    if not isinstance(formula, str):
        return None
    m = _HYPERLINK_CODE_RE.search(formula.strip())
    return m.group(1) if m else None

def _smart_to_float(x) -> Optional[float]:
    """Converte string com pt-BR/EN para float. Mantém float intacto; '-' ou vazio -> None."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s in ("", "-"):
        return None
    # remove lixo, mas preserva '.' e ',' para decidir semântica
    s = re.sub(r"[^0-9\.,-]", "", s)
    has_dot, has_comma = "." in s, "," in s
    if has_dot and has_comma:
        s = s.replace(".", "").replace(",", ".")  # . = milhar, , = decimal
    elif has_comma and not has_dot:
        s = s.replace(",", ".")                   # só vírgula -> decimal
    # só ponto -> já está em EN
    try:
        return float(s)
    except ValueError:
        return None

# ----------------- loader principal -----------------

def load_sinapi_ccd_pr(path: str, cidade: str = "CURITIBA") -> CanonDict:
    """
    Lê a aba CCD do SINAPI e retorna Dict[codigo, Item] usando a coluna ('PR', cidade) como CUSTO.
    Extrai código da fórmula HYPERLINK; lê código/descrição/custo **da mesma linha** (openpyxl),
    iniciando no primeiro código numérico para evitar deslocamentos de observações no topo.
    """
    # 1) Use pandas só para detectar header e localizar índices de colunas
    dfm = pd.read_excel(path, sheet_name="CCD", header=[3, 4])
    header_row = 4                      # linha com rótulos "Grupo / Código / Descrição" (idx pandas)
    probe = dfm.iloc[header_row]

    def pick_first(*starts: str):
        for col in dfm.columns:
            if any(_norm(str(probe[col])).startswith(_norm(s)) for s in starts):
                return col
        return None

    col_grupo  = pick_first("grupo")
    col_codigo = pick_first("codigo", "código")
    col_desc   = pick_first("descricao", "descrição")
    if not all([col_grupo, col_codigo, col_desc]):
        raise RuntimeError(f"[SINAPI CCD] Não encontrei colunas básicas. "
                           f"grupo={col_grupo}, codigo={col_codigo}, desc={col_desc}")

    # custo do PR: ('PR', cidade) — a subcoluna .1 é %AS; evitamos ela
    pr_custo_col = None
    for a, b in dfm.columns:
        if _norm(a) == "pr" and _norm(b) == _norm(cidade):
            pr_custo_col = (a, b)
            break
    if pr_custo_col is None:
        candidates = [c for c in dfm.columns
                      if c[0] == "PR" and _norm(c[1]).startswith(_norm(cidade)) and not str(c[1]).endswith(".1")]
        if not candidates:
            raise RuntimeError(f"[SINAPI CCD] Coluna de custo PR/{cidade} não encontrada.")
        pr_custo_col = candidates[0]

    # Índices de coluna (1-based no Excel) conforme a ordem do pandas
    x_col_codigo = list(dfm.columns).index(col_codigo) + 1
    x_col_desc   = list(dfm.columns).index(col_desc)   + 1
    x_col_custo  = list(dfm.columns).index(pr_custo_col) + 1

    # 2) Leia diretamente do Excel (openpyxl) para manter as três colunas alinhadas por linha
    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb["CCD"]

    # a primeira linha de dados (em Excel) é header_row+2
    start_row_excel = (header_row + 1) + 1

    # Avance até encontrar o primeiro código numérico (evita bloco de observações)
    r = start_row_excel
    max_r = ws.max_row
    first_data_row = None
    while r <= max_r:
        v = ws.cell(row=r, column=x_col_codigo).value
        code = _extract_code_from_formula(v) if isinstance(v, str) and v.startswith("=") \
               else (str(v).strip() if v is not None else None)
        if isinstance(code, str) and _DIGIT_CODE_RE.fullmatch(code):
            first_data_row = r
            break
        r += 1
    if first_data_row is None:
        raise RuntimeError("[SINAPI CCD] Não encontrei nenhum código numérico na CCD.")

    # 3) Varra da primeira linha válida até o fim, coletando código/descrição/custo
    out: CanonDict = {}
    dup = 0
    for row in ws.iter_rows(min_row=first_data_row, max_row=max_r,
                            min_col=min(x_col_codigo, x_col_desc, x_col_custo),
                            max_col=max(x_col_codigo, x_col_desc, x_col_custo),
                            values_only=False):
        # A linha inteira veio; pegue só as 3 células relevantes
        c_code  = row[x_col_codigo  - min(x_col_codigo, x_col_desc, x_col_custo)]
        c_desc  = row[x_col_desc    - min(x_col_codigo, x_col_desc, x_col_custo)]
        c_custo = row[x_col_custo   - min(x_col_codigo, x_col_desc, x_col_custo)]

        # código
        vcode = c_code.value
        code = _extract_code_from_formula(vcode) if isinstance(vcode, str) and vcode.startswith("=") \
               else (str(vcode).strip() if vcode is not None else None)
        if not (isinstance(code, str) and _DIGIT_CODE_RE.fullmatch(code)):
            # acabou a sequência de dados (normalmente após o último bloco)
            continue

        # descrição
        desc = c_desc.value
        desc = "" if desc is None else str(desc).strip()

        # custo PR
        custo = _smart_to_float(c_custo.value)

        # alguns finais de bloco trazem custo vazio; mantemos mas com 0.0
        if custo is None:
            custo = 0.0

        item: Item = {
            "codigo": norm_code(code),
            "descricao": desc,
            "valor_unit": float(custo),
            "fonte": "SINAPI",
        }
        if item["codigo"] in out:
            dup += 1
        out[item["codigo"]] = item

    if dup:
        logger.warning("SINAPI CCD PR: %d código(s) duplicado(s); mantendo o último.", dup)

    return out
