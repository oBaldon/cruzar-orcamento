# src/cruzar_orcamento/exporters/excel.py
from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd

def _autofit_columns(ws) -> None:
    """Ajusta largura das colunas com base no conteúdo (openpyxl worksheet)."""
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = cell.value
            val_str = str(val) if val is not None else ""
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 80)

def export_cruzamento_excel(
    cruzado: List[Dict[str, Any]],
    divergencias: List[Dict[str, Any]],
    path: str | Path,
    *,
    round_decimals: int = 2,
    number_format_currency: str = '#,##0.00',
    number_format_percent: str = '0.00%',
) -> Path:
    """
    Gera um arquivo Excel com:
      - aba 'cruzado' (todas as linhas, com dif_abs e dif_rel calculados)
      - aba 'divergencias' (apenas divergências, já com dif_abs e dif_rel do processor)

    Parâmetros de formatação podem ser ajustados conforme necessidade.
    Retorna o Path do arquivo gerado.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # --- montar DataFrame do cruzado e calcular difs ---
    df_cruz = pd.DataFrame(cruzado)
    if not df_cruz.empty:
        # garantir colunas
        for col in ["a_valor", "b_valor"]:
            if col not in df_cruz.columns:
                df_cruz[col] = None

        # calcular dif_abs e dif_rel
        def _dif_abs(a, b):
            if a is None or b is None:
                return None
            try:
                return abs(float(a) - float(b))
            except Exception:
                return None

        def _dif_rel(a, b):
            if a is None or b in (None, 0):
                return None
            try:
                return abs(float(a) - float(b)) / float(b) if float(b) != 0 else None
            except Exception:
                return None

        df_cruz["dif_abs"] = [ _dif_abs(a, b) for a, b in zip(df_cruz["a_valor"], df_cruz["b_valor"]) ]
        df_cruz["dif_rel"] = [ _dif_rel(a, b) for a, b in zip(df_cruz["a_valor"], df_cruz["b_valor"]) ]

        # arredondar valores numéricos (sem estragar None)
        for col in ["a_valor", "b_valor", "dif_abs"]:
            df_cruz[col] = pd.to_numeric(df_cruz[col], errors="coerce").round(round_decimals)
        # dif_rel em fração (0.0123 = 1.23%); não arredonda aqui, aplica formato no Excel
    else:
        df_cruz = pd.DataFrame(columns=[
            "codigo","a_banco","a_desc","a_valor","b_desc","b_valor","match","dif_abs","dif_rel"
        ])

    # --- DataFrame de divergências ---
    df_div = pd.DataFrame(divergencias) if divergencias else pd.DataFrame(
        columns=["codigo","motivos","dif_abs","dif_rel"]
    )
    if not df_div.empty:
        if "dif_abs" in df_div.columns:
            df_div["dif_abs"] = pd.to_numeric(df_div["dif_abs"], errors="coerce").round(round_decimals)
        # dif_rel permanece fração para receber formato percentual no Excel

    # --- escrever Excel com formatação ---
    with pd.ExcelWriter(path, engine="openpyxl") as xlw:
        df_cruz.to_excel(xlw, sheet_name="cruzado", index=False)
        df_div.to_excel(xlw, sheet_name="divergencias", index=False)

        wb = xlw.book

        # Formatar 'cruzado'
        ws_c = wb["cruzado"]
        _autofit_columns(ws_c)
        # localizar índices das colunas para aplicar formato
        headers_c = [c.value for c in next(ws_c.iter_rows(min_row=1, max_row=1))]
        def col_idx(hdr: str) -> Optional[int]:
            try:
                return headers_c.index(hdr) + 1
            except ValueError:
                return None

        c_a = col_idx("a_valor")
        c_b = col_idx("b_valor")
        c_da = col_idx("dif_abs")
        c_dr = col_idx("dif_rel")

        from openpyxl.styles import numbers
        for r in ws_c.iter_rows(min_row=2):
            if c_a: r[c_a-1].number_format = number_format_currency
            if c_b: r[c_b-1].number_format = number_format_currency
            if c_da: r[c_da-1].number_format = number_format_currency
            if c_dr: r[c_dr-1].number_format = number_format_percent

        # Formatar 'divergencias'
        ws_d = wb["divergencias"]
        _autofit_columns(ws_d)
        headers_d = [c.value for c in next(ws_d.iter_rows(min_row=1, max_row=1))]
        def d_idx(hdr: str) -> Optional[int]:
            try:
                return headers_d.index(hdr) + 1
            except ValueError:
                return None

        d_da = d_idx("dif_abs")
        d_dr = d_idx("dif_rel")
        for r in ws_d.iter_rows(min_row=2):
            if d_da: r[d_da-1].number_format = number_format_currency
            if d_dr: r[d_dr-1].number_format = number_format_percent

    return path
